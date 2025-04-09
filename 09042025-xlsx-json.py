import openpyxl
import json


input_file = 'C:\\Users\\fdomingues\\JCDECAUX\\BR - Dados - Documentos\\01. Soluções\\MAPA OOH Live\\01. Projetos\\01. Praças\\01. Atualização Inventario (Todas as Praças)\\database_frontend.xlsx'
output_file = 'C:\\Users\\fdomingues\\JCDECAUX\\BR - Dados - Documentos\\01. Soluções\\MAPA OOH Live\\01. Projetos\\01. Praças\\01. Atualização Inventario (Todas as Praças)\\database_frontend.json'

workbook = openpyxl.load_workbook(input_file)
sheet = workbook.active

headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))

data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_dict = {}
    for header, value in zip(headers, row):
        row_dict[header] = value if value is not None else ""
    data.append(row_dict)

with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Arquivo convertido com sucesso para {output_file}")